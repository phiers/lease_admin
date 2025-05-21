from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import os
import sys
from datetime import datetime, timedelta
from calendar import monthrange
import shutil


def main():
    process = start_program()
    run_process(process)


def start_program():
    """Prompt user for process to run"""
    while True:
        try:
            process = input(
                """What would you like to do?
            1 - Process Lucernex files
            2 - Create initial analysis file
            3 - Create final analysis file
            4 - Create csv file
            5 - Exit this program
                 \n Enter the appropriate number: """
            )
            if int(process) in [1, 2, 3, 4, 5, 6]:
                return int(process)
        except:
            print("\nENTER A NUMBER BETWEEN 1 AND 5! \n")
            continue


def run_process(process):
    """Define which functions to run for each process"""
    if process == 1:
        directories = [
            "1_Lx_files",
            "2_lease_files",
            "3_equip_files",
            "4_input_files",
            "5_output_files",
        ]
        check_dir_structure(directories)
        rename_and_move_files("1_Lx_files", "2_lease_files")
        separate_express_file("2_lease_files", "express.xlsx")
        separate_hatch_file("2_lease_files", "janieandjack.xlsx")
    if process == 2:
        date = get_date()
        add_data = process_additional_invoice_items(
            date, Path.cwd().joinpath("4_input_files", "additional_invoice_items.csv")
        )
        results = process_files_and_create_dict("2_lease_files", add_data, date)
        create_initial_analysis(results, date)
    if process == 3:
        date = get_date()
        m, _, y = date.split("/")
        path = Path.cwd().joinpath(
            "5_output_files", f"{m}_{y}_initial_invoice_analysis.xlsx"
        )
        create_final_analysis_files(path, date)
    if process == 4:
        date = get_date()
        m, _, y = date.split("/")
        create_csv_from_analysis_file(
            f"5_output_files/{m}_{y}_final_invoice_analysis.xlsx"
        )
    if process == 5:
        sys.exit()


def check_dir_structure(paths):
    for path in paths:
        path = Path.cwd().joinpath(path)
        folder = str(path).split("/")[-1]
        if os.path.isdir(path) == False:
            sys.exit(
                f"ERROR: The {folder} folder does not exist. Please add one as a subfolder in the Lease_Admin folder."
            )


def rename_and_move_files(directory, target_dir):
    lease_file_count = 0
    try:
        for file in os.scandir(directory):
            if file.name.split(".")[1] == "xlsx":
                wb = load_workbook(file)
                ws = wb.active
                sheet_identifier_list = ws["A1"].value.split(" ")
                if "Equipment" in sheet_identifier_list:
                    name = sheet_identifier_list[-1].split(".")[0] + "_equipment.xlsx"
                    wb.save(Path.cwd().joinpath("3_equip_files", name))
                else:
                    name = sheet_identifier_list[-1].split(".")[0] + ".xlsx"
                    wb.save(Path.cwd().joinpath(target_dir, name))
                    lease_file_count += 1
    except FileNotFoundError:
        print(f"ERROR: the '{directory}' folder does not exist or is empty.")

    if lease_file_count == 0:
        print(f"There are no files in the {directory} folder!")
    else:
        print(
            f"{lease_file_count} lease files were created and saved in the lease_files folder"
        )


def separate_express_file(dir, file):
    """Separate homage, bonobos and express (all in express file when sent from Lucernex)"""
    df = pd.read_excel(Path.cwd().joinpath(dir, file), header=1)
    cols = df.columns
    # find homage leases and make new file
    df_homage = df[df["Contract Name"].str.startswith("Hom")]
    df_homage.to_excel(Path.cwd().joinpath(dir, "homage.xlsx"), columns=cols)
    # find Bonobos leases and make new file
    df_bonobos = df[df["Contract ID"].str.startswith("05")]
    df_bonobos.to_excel(Path.cwd().joinpath(dir, "bonobos.xlsx"), columns=cols)
    # remove homage from df and save separate express file
    df = df[~df["Contract Name"].str.startswith("Hom")]
    df_express = df[~df["Contract ID"].str.startswith("05")]
    df_express.to_excel(Path.cwd().joinpath(dir, "express_only.xlsx"), columns=cols)
    # delete the original file
    os.remove(f"{dir}/{file}")

def separate_hatch_file(dir, file):
    """Separate hatch from janie and jack file"""
    df = pd.read_excel(Path.cwd().joinpath(dir, file), header=1)
    cols = df.columns
    # find hatch leases and make new file
    df_hatch = df[df["Contract ID"].astype(str).str.startswith(("5", "9"))]
    df_hatch.to_excel(Path.cwd().joinpath(dir, "hatch.xlsx"), columns=cols)
    # remove hatch from df and save separate janie and jack file
    df = df[~df["Contract Name"].str.startswith("Hom")]
    df_jnj = df[~df["Contract ID"].astype(str).str.startswith(("5", "9"))]
    df_jnj.to_excel(Path.cwd().joinpath(dir, "janie_and_jack_only.xlsx"), columns=cols)
    # delete the original file
    os.remove(f"{dir}/{file}")


def get_date():
    """returns previous monthend date in mm/dd/yyyy format"""
    t = datetime.today().date() + timedelta(days=5)
    if t.month != 1:
        prev_month = t.month - 1
        year = t.year
    else:
        # if month is January, make previous month December
        prev_month = 12
        year = t.year - 1
    # get max days in each month
    days = monthrange(t.year, prev_month)[1]
    # format months 1 thru 9 to include leading zero
    if prev_month < 10:
        prev_month = f"0{prev_month}"

    return f"{prev_month}/{days}/{year}"


def create_cust_name_dict():
    """Helper function to create customer names dictionary (file names xref to NetSuite names)"""
    try:
        return (
            pd.read_csv("4_input_files/customer_names.csv", header=None, index_col=0)
            .squeeze()
            .to_dict()
        )
    except FileNotFoundError:
        sys.exit("customer_names.csv file must be in input_files directory")


def process_additional_invoice_items(date, file_path):
    """
    Reads additional_invoice_items.csv and returns lists used in creating initial invoice analysis
    """
    df = pd.read_csv(file_path, usecols=[0, 1, 2]).dropna(thresh=2)
    customer_name_dict = create_cust_name_dict()
    descriptions = [x for x in df["Description"]]
    quantities = [int(x) for x in df["Quantity"]]
    dates = [date for x in range(df.shape[0])]
    clients = []
    for name in df["Customer_File_Name"]:
        try:
            clients.append(customer_name_dict[name])
        except:
            clients.append(name)
            print(
                f"WARNING: customer {name} not in customer_names.csv file. Please add name to the file and rerun process 2"
            )

    lx_type_codes = [str(x) + "_" + y for x, y in zip(clients, descriptions)]

    return clients, dates, lx_type_codes, quantities, descriptions


def process_files_and_create_dict(directory, addl_invoice_items, date):
    """
    Returns a dictionary of data necessary to create this month's items and quantities on the initial invoice analysis, 
    including data from the entered in additional_invoice_items.csv
    """
    customer_name_dict = create_cust_name_dict()
    # create lists from additional_invoice_items() return values
    clients = addl_invoice_items[0]
    descriptions = addl_invoice_items[4]
    lx_type_codes = addl_invoice_items[2]
    quantities = addl_invoice_items[3]
    dates = addl_invoice_items[1]
    # blank dictionary for return
    results_dict = {}
    # create rest of results from processed lease count files (TODO: move to separate function?)
    for file in os.scandir(directory):
        if file.name.split(".")[1] == "xlsx":
            try:
                client_name = customer_name_dict[file.name.split(".")[0]]
            except KeyError:
                print(
                    f"ERROR! The customer name does not exist on customer_names.csv for {file.name}. If the name is not added, this file will not be processed."
                )

            # the split files start at row 0, not 1
            header = 0 if file.name in {"homage.xlsx", "express_only.xlsx", "bonobos.xlsx", "hatch.xlsx", "janie_and_jack_only.xlsx"} else 1
            
            df = pd.read_excel(file, header=header)
            try:
                # distinguish between domestic and international for Tory
                if file.name == "tory.xlsx" or file.name == "asics.xlsx":
                    data = df.loc[:, ["Lease Status", "Region"]]
                    for key, value in data.value_counts().items():
                        if key[1] == "North America" or key[1] == "United States":
                            clients.append(client_name)
                            descriptions.append(f"{key[0]} - Domestic")
                            lx_type_codes.append(
                                f"{client_name}_{key[0]} - Domestic"
                            )
                            quantities.append(value)
                            dates.append(date)
                        else:
                            clients.append(client_name)
                            descriptions.append(f"{key[0]} - International")
                            lx_type_codes.append(
                                f"{client_name}_{key[0]} - International - {key[1]}"
                            )
                            quantities.append(value)
                            dates.append(date)
                else:
                    data = df.loc[:, "Lease Status"]
                    for key, value in data.value_counts().items():
                        clients.append(client_name)
                        descriptions.append(key)
                        lx_type_codes.append(f"{client_name}_{key}")
                        quantities.append(value)
                        dates.append(date)

            # throw error if file is not processsed (i.e., doesn't have "Lease Status" column)
            except KeyError:
                print(KeyError.with_traceback)
                print(f"ERROR! {file.name} was not processed or had no data")

    results_dict["Customer"] = clients
    results_dict["Date"] = dates
    results_dict["Lx_Type"] = descriptions
    results_dict["Lx_Type_Code"] = lx_type_codes
    results_dict["Quantity"] = quantities
    
    return results_dict


def create_price_and_description_df():
    """Creates dataframe dataframe with prices and descriptions from type_desc_price_matrix.xlsx"""
    try:
        return pd.read_excel(
            Path.cwd().joinpath("4_input_files", "type_desc_price_matrix.xlsx"), usecols=[0, 1, 2]
        )
    except FileNotFoundError:
        print(
            "ERROR: A file named 'type_desc_price_matrix.csv' must be in the '4_input_files' directory"
        )


# helper function to create dataframe for last month's data from excel
def create_lm_df():
    """returns dataframe with last month's data from lm_invoice_analysis.xlsx"""
    try:
        df = pd.read_excel(
            # TODO: change filename to grab prior month
            Path.cwd().joinpath("4_input_files", "lm_invoice_analysis.xlsx"),
            usecols=["Lx_Type_Code", "Quantity", "Price"],
        )
    except FileNotFoundError:
        print(
            "A file named 'lm_invoice_analysis.xlsx' must be in the input files directory"
        )
    else:
        # rename columns
        df.columns = ["Lx_Type_Code", "LM_Quantity", "LM_Price"]
        # get rid of rows without quantity
        df = df.fillna(0)
        df = df[df["LM_Quantity"] > 0]
        # drop totals row
        df = df[:-1]

        return df


def create_initial_analysis(dic, date):
    # create df with monthly invoice data
    df_monthly_data = pd.DataFrame.from_dict(dic)
    # rename price and description dataframe
    df_price_desc = create_price_and_description_df()
    # create df with last month's data
    df_lm = create_lm_df()
    initial_combined_df = pd.merge(
        df_monthly_data,
        df_lm,
        how="outer",
        on="Lx_Type_Code",
    )

    combined_df = pd.merge(
        initial_combined_df, df_price_desc, how="left", on="Lx_Type_Code"
    )
    # sort rows
    combined_df = combined_df.sort_values(["Lx_Type_Code", "Invoice_Description"])

    # make NaN values zero for calcs
    combined_df[["Price", "Quantity", "LM_Quantity", "LM_Price"]] = combined_df[
        ["Price", "Quantity", "LM_Quantity", "LM_Price"]
    ].fillna(value=0)
    # create total and difference columns vs lm
    total = combined_df["Quantity"] * combined_df["Price"]
    lm_total = combined_df["LM_Price"] * combined_df["LM_Quantity"]
    combined_df.insert(0, "Total", total)
    combined_df.insert(1, "LM_Total", lm_total)
    qnty_vs_lm = combined_df["Quantity"] - combined_df["LM_Quantity"]
    price_vs_lm = combined_df["Price"] - combined_df["LM_Price"]
    total_vs_lm = combined_df["Total"] - combined_df["LM_Total"]
    combined_df.insert(2, "Qnty_vs_LM", qnty_vs_lm)
    combined_df.insert(3, "Price_vs_LM", price_vs_lm)
    combined_df.insert(4, "Total_vs_LM", total_vs_lm)

    # sort columns in better order
    col_order = [
        "Date",
        "Customer",
        "Lx_Type",
        "Lx_Type_Code",
        "Invoice_Description",
        "Quantity",
        "Price",
        "Total",
        "LM_Quantity",
        "LM_Price",
        "LM_Total",
        "Qnty_vs_LM",
        "Price_vs_LM",
        "Total_vs_LM",
    ]
    combined_df = combined_df[col_order]
    # total numeric columns
    columns_to_total = [
        "Quantity",
        "Total",
        "LM_Quantity",
        "LM_Total",
        "Qnty_vs_LM",
        "Total_vs_LM",
    ]
    combined_df.loc["Totals"] = combined_df.loc[:, columns_to_total].sum(axis=0)
    # create a name to save the file under
    month, _, year = date.split("/")
    save_file_path = Path.cwd().joinpath(
        "5_output_files", f"{month}_{year}_initial_invoice_analysis.xlsx"
    )
    combined_df.to_excel(save_file_path, index=False)

    print(f"initial analysis file for {date} created")


def create_final_analysis_files(file_path, date):
    initial_df = pd.read_excel(file_path).round(2)
    sum_df = (
        initial_df.groupby(["Customer", "Invoice_Description"])
        .agg(
            {
                "Quantity": "sum",
                "Price": "mean",
                "Total": "sum",
                "LM_Quantity": "sum",
                "LM_Price": "mean",
                "LM_Total": "sum",
                "Qnty_vs_LM": "sum",
                "Price_vs_LM": "sum",
                "Total_vs_LM": "sum",
            }
        )
        .round(2)
        .reset_index()
    )

    cust_sum_df = (
        sum_df.groupby("Customer")
        .agg({"Total": "sum", "LM_Total": "sum", "Total_vs_LM": "sum"})
        .round(2)
        .reset_index()
    )
    # add totals to columns
    cust_sum_df.loc["Totals"] = cust_sum_df.sum(axis=0, numeric_only=True)
    sum_df.loc["Totals"] = sum_df.sum(axis=0, numeric_only=True)

    m, _, y = date.split("/")
    save_file_name = f"{m}_{y}_final_invoice_analysis.xlsx"
    # write to excel as three separate sheets
    with pd.ExcelWriter(
        Path.cwd().joinpath("5_output_files", save_file_name)
    ) as writer:
        cust_sum_df.to_excel(writer, sheet_name="cust_summary", index=False)
        sum_df.to_excel(writer, sheet_name="summary", index=False)
        initial_df.to_excel(writer, sheet_name="detail", index=False)

    print("Final analysis files created")


def create_csv_from_analysis_file(f):
    full_date = get_date()
    month, _, year = full_date.split("/")
    df = pd.read_excel(
        f,
        sheet_name="summary",
        usecols=["Customer", "Invoice_Description", "Quantity", "Price"],
    )
    cust_str = df["Customer"].str.replace(" ", "_")
    cust_str = cust_str.replace(",", "_").replace(".", "_").str[0:5]
    ext_id = year[2:] + month + cust_str
    df.insert(0, "External ID", ext_id)
    df.insert(1, "Date", full_date)
    df.insert(4, "Memo", "Lease Admin Services")
    df = df[:-1]
    df = df[df["Quantity"] != 0]

    output_path = Path("5_output_files") / f"{month}{year[2:]}_invoice_upload.csv"
    df.to_csv(output_path, index=False)

if __name__ == "__main__":
    main()
